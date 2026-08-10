# -*- coding: utf-8 -*-
"""2차 지원자 12명 → 학과별 엑셀 분할 + 교수님별 배포 폴더(엑셀·PDF) 구성.

1차 구조를 그대로 따른다:
  학과별 지원자 현황(엑셀)\2차\엑셀파일 모음\(과정)학과_N명.xlsx
  학과별 지원자 현황(엑셀)\2차\교수님별 배포\학과(교수)\...
  학과별 지원자 서류 일체(PDF)\2차\교수님별 배포\학과(교수)\...
  · 파일이 2개 이상이면 `학과(교수).zip` 으로 묶는다 (1차 관례)
  · 엑셀은 **데이터가 있는 학력 블록만** 남기고 K열부터 재배치 (1차 관례)
"""
import os, re, shutil, zipfile
import openpyxl

BASE = r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기'
DIST = os.path.join(BASE, r'지원서류\학과 평가\학과별 배포')
SRC = os.path.join(BASE, '(2차)2026학년도 후기 대학원 지원자 현황_12명.xlsx')
TPL_DIR = os.path.join(DIST, r'학과별 지원자 현황(엑셀)\1차\엑셀파일 모음')

XL_2CHA = os.path.join(DIST, r'학과별 지원자 현황(엑셀)\2차')
XL_POOL = os.path.join(XL_2CHA, '엑셀파일 모음')
XL_PROF = os.path.join(XL_2CHA, '교수님별 배포')
PDF_2CHA = os.path.join(DIST, r'학과별 지원자 서류 일체(PDF)\2차')
PDF_POOL = os.path.join(PDF_2CHA, 'PDF 파일 모음')
PDF_PROF = os.path.join(PDF_2CHA, '교수님별 배포')

# 학과 → 전공주임교수 (2025-2학기 주임교수 현황(2026.01.01. 기준) + 1차 배포 실적)
PROF = {
    '간호학과': '강광순', '컴퓨터공학과': '이익훈', '법학과': '남윤경',
    '임상상담심리학과': '박주영', '미래융합기술공학과': '최재혁',
    '물류유통경영학과': '김현종', '시각영상디자인': '김한성',   # 디자인학과 주임교수
}

# 레이아웃별 템플릿 (1차 파일 재사용 — 서식·테두리 100% 보존)
#   키: (전문학사 포함, 석사 포함) / 값: (템플릿 파일, 템플릿 데이터 행수)
TPL = {
    (True,  True):  ('(박사)경영학과_1명.xlsx', 1),
    (True,  False): ('(석사)법학과_5명.xlsx', 5),
    (False, True):  ('(박사)한국어문화교육콘텐츠학과_3명.xlsx', 3),
    (False, False): ('(석사)물류유통경영학과_3명.xlsx', 3),
}

# 원본 열 인덱스(1-based)
COMMON = list(range(1, 11))          # A~J
COL_BLK = list(range(11, 16))        # K~O 전문학사(5)
UNI_BLK = list(range(16, 20))        # P~S 학사(4)
MA_BLK = list(range(20, 24))         # T~W 석사(4)

ws_src = openpyxl.load_workbook(SRC).active
recs = []
for r in range(3, ws_src.max_row + 1):
    if ws_src.cell(r, 1).value is None:
        continue
    recs.append([ws_src.cell(r, c).value for c in range(1, 24)])

# 그룹핑: (과정 기본형, 지원학과) — '석사편입'은 석사 그룹에 합류(PDF 명명과 동일)
groups = {}
for rec in recs:
    base = '박사' if str(rec[1]).startswith('박사') else '석사'
    groups.setdefault((base, str(rec[4]).strip()), []).append(rec)

os.makedirs(XL_POOL, exist_ok=True)
os.makedirs(XL_PROF, exist_ok=True)
os.makedirs(PDF_PROF, exist_ok=True)

made = []
for (course, dept), rows in groups.items():
    has_col = any(r[10] for r in rows)
    has_ma = any(r[19] for r in rows)
    tpl_name, tpl_rows = TPL[(has_col, has_ma)]
    assert len(rows) <= tpl_rows, f'{dept}: 템플릿 행 부족 ({len(rows)}>{tpl_rows})'

    out = os.path.join(XL_POOL, f'({course}){dept}_{len(rows)}명.xlsx')
    shutil.copyfile(os.path.join(TPL_DIR, tpl_name), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active

    src_cols = COMMON + (COL_BLK if has_col else []) + UNI_BLK + (MA_BLK if has_ma else [])
    for i, rec in enumerate(rows):
        for j, sc in enumerate(src_cols):
            ws.cell(3 + i, j + 1).value = (i + 1) if sc == 1 else rec[sc - 1]

    last = 2 + len(rows)
    if tpl_rows > len(rows):
        ws.delete_rows(last + 1, tpl_rows - len(rows))
    ws.auto_filter.ref = f'A2:{openpyxl.utils.get_column_letter(len(src_cols))}{last}'
    wb.save(out)
    made.append((course, dept, [str(r[2]) for r in rows], os.path.basename(out), ws.dimensions))

print('=== 엑셀 분할 ===')
for c, d, names, f, dim in sorted(made, key=lambda x: (x[0], x[1])):
    print(f'  {f:<34} {dim:<8} {", ".join(names)}')

# ── 교수님별 배포 (엑셀·PDF 공통 로직) ─────────────────────────────────
def distribute(files, dest_root, dept, prof):
    """파일 1개면 그대로, 2개 이상이면 학과(교수).zip 으로 묶어 배치."""
    folder = os.path.join(dest_root, f'{dept}({prof})')
    os.makedirs(folder, exist_ok=True)
    if len(files) == 1:
        shutil.copyfile(files[0], os.path.join(folder, os.path.basename(files[0])))
        return os.path.basename(files[0])
    zp = os.path.join(folder, f'{dept}({prof}).zip')
    with zipfile.ZipFile(zp, 'w', zipfile.ZIP_DEFLATED) as z:
        for f in sorted(files):
            z.write(f, os.path.basename(f))
    return os.path.basename(zp)


by_dept = {}
for course, dept, names, fname, _ in made:
    by_dept.setdefault(dept, {'xlsx': [], 'names': []})
    by_dept[dept]['xlsx'].append(os.path.join(XL_POOL, fname))
    by_dept[dept]['names'] += names

pdfs = {f: os.path.join(PDF_POOL, f) for f in os.listdir(PDF_POOL) if f.lower().endswith('.pdf')}

print('\n=== 교수님별 배포 ===')
used_pdf = set()
for dept, info in sorted(by_dept.items()):
    prof = PROF[dept]
    x = distribute(sorted(info['xlsx']), XL_PROF, dept, prof)
    mine = [p for f, p in pdfs.items() if f'){dept}_' in f]
    used_pdf |= set(mine)
    p = distribute(sorted(mine), PDF_PROF, dept, prof) if mine else '(PDF 없음)'
    print(f'  {dept}({prof}) — 지원자 {len(info["names"])}명 {info["names"]}')
    print(f'      엑셀: {x}')
    print(f'      PDF : {p} ({len(mine)}건)')

left = set(pdfs.values()) - used_pdf
if left:
    print('\n⚠ 배포되지 않은 PDF:', [os.path.basename(p) for p in sorted(left)])
else:
    print(f'\nPDF {len(pdfs)}건 전량 배포 완료')
