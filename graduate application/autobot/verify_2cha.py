# -*- coding: utf-8 -*-
"""분할본 검증: ①12명 전원 1회씩 존재 ②원본과 전 항목 일치 ③서식이 1차 템플릿과 동일 ④PDF 대응."""
import os, glob, zipfile
import openpyxl

BASE = r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기'
DIST = os.path.join(BASE, r'지원서류\학과 평가\학과별 배포')
SRC = os.path.join(BASE, '(2차)2026학년도 후기 대학원 지원자 현황_12명.xlsx')
POOL = os.path.join(DIST, r'학과별 지원자 현황(엑셀)\2차\엑셀파일 모음')
TPL_DIR = os.path.join(DIST, r'학과별 지원자 현황(엑셀)\1차\엑셀파일 모음')
PDF_POOL = os.path.join(DIST, r'학과별 지원자 서류 일체(PDF)\2차\PDF 파일 모음')
XL_PROF = os.path.join(DIST, r'학과별 지원자 현황(엑셀)\2차\교수님별 배포')
PDF_PROF = os.path.join(DIST, r'학과별 지원자 서류 일체(PDF)\2차\교수님별 배포')

ws = openpyxl.load_workbook(SRC).active
src = {}
for r in range(3, ws.max_row + 1):
    if ws.cell(r, 1).value is None:
        continue
    v = [ws.cell(r, c).value for c in range(1, 24)]
    src[str(v[2])] = v

HDR_BLK = {'전문학사': 5, '학사': 4, '석사': 4}
bad = 0
seen = []
print('=== ① 데이터 대조 ===')
for p in sorted(glob.glob(os.path.join(POOL, '*.xlsx'))):
    w = openpyxl.load_workbook(p).active
    # 헤더에서 블록 구성 파악
    blocks, c = [], 11
    while c <= w.max_column:
        h = w.cell(1, c).value
        if h in HDR_BLK:
            blocks.append(h); c += HDR_BLK[h]
        else:
            c += 1
    idx = list(range(1, 11))
    for b in blocks:
        idx += {'전문학사': list(range(11, 16)), '학사': list(range(16, 20)), '석사': list(range(20, 24))}[b]

    n = 0
    for r in range(3, w.max_row + 1):
        if w.cell(r, 1).value is None:
            continue
        n += 1
        name = str(w.cell(r, 3).value)
        seen.append(name)
        exp = src[name]
        diff = []
        for j, sc in enumerate(idx):
            got = w.cell(r, j + 1).value
            want = n if sc == 1 else exp[sc - 1]
            if (got or None) != (want or None):
                diff.append(f'{openpyxl.utils.get_column_letter(j+1)}열 원본={want!r} 분할={got!r}')
        # 제외된 블록에 데이터가 없었는지 확인(정보 유실 방지)
        for b, cols in (('전문학사', range(11, 16)), ('석사', range(20, 24))):
            if b not in blocks and any(exp[k - 1] for k in cols):
                diff.append(f'{b} 블록 누락인데 원본에 데이터 있음!')
        if diff:
            bad += 1
            print(f'  ❌ {os.path.basename(p)} / {name}: ' + '; '.join(diff))
    print(f'  ✅ {os.path.basename(p):<32} {n}명 블록={blocks}')

print(f'\n=== ② 인원 대조 ===')
print(f'  원본 {len(src)}명 / 분할 합계 {len(seen)}명 / 중복 {len(seen)-len(set(seen))}건 / 누락 {sorted(set(src)-set(seen))}')

print('=== ③ 서식(1차 템플릿 대비) ===')
def sig(w):
    a = w[ 'A1' ]; d = w['A3']
    return (w.title, sorted(str(m) for m in w.merged_cells.ranges if ':' in str(m) and str(m)[1] in '12'),
            w.freeze_panes, {k: round(v.width, 1) for k, v in w.column_dimensions.items() if v.width},
            a.font.bold, a.fill.fill_type, a.alignment.horizontal, a.border.left.style, d.border.left.style)
TPLMAP = {'A1:W': '(박사)경영학과_1명.xlsx', 'A1:S': '(석사)법학과_5명.xlsx',
          'A1:R': '(박사)한국어문화교육콘텐츠학과_3명.xlsx', 'A1:N': '(석사)물류유통경영학과_3명.xlsx'}
for p in sorted(glob.glob(os.path.join(POOL, '*.xlsx'))):
    w = openpyxl.load_workbook(p).active
    key = 'A1:' + w.dimensions.split(':')[1].rstrip('0123456789')
    t = openpyxl.load_workbook(os.path.join(TPL_DIR, TPLMAP[key])).active
    ok = sig(w) == sig(t)
    print(f'  {"✅" if ok else "❌"} {os.path.basename(p):<32} ← {TPLMAP[key]}')
    if not ok:
        bad += 1

print('=== ④ 교수님별 배포 ===')
for root, label in ((XL_PROF, '엑셀'), (PDF_PROF, 'PDF')):
    for d in sorted(os.listdir(root)):
        fs = os.listdir(os.path.join(root, d))
        detail = []
        for f in fs:
            if f.endswith('.zip'):
                with zipfile.ZipFile(os.path.join(root, d, f)) as z:
                    detail.append(f'{f} → {len(z.namelist())}건 {z.namelist()}')
            else:
                detail.append(f)
        print(f'  [{label}] {d}: ' + ' | '.join(detail))

pdfs = [f for f in os.listdir(PDF_POOL) if f.endswith('.pdf')]
print(f'\n=== 결과: 오류 {bad}건 / PDF 원본 {len(pdfs)}건 / 지원자 {len(src)}명 ===')
