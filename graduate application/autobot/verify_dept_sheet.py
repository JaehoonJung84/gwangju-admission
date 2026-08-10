# -*- coding: utf-8 -*-
"""생성본 ↔ 원본 엑셀 전건 대조 (이름·영문·학과·국적·성별·등록번호 생년월일·TOPIK·학력 3블록)."""
import json, re, datetime
import openpyxl

SRC = r'C:\projects\graduate application\autobot\rows2.json'
OUT = (r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기'
       r'\(2차)2026학년도 후기 대학원 지원자 현황_11명.xlsx')

rows = [r['c'] for r in json.load(open(SRC, encoding='utf-8-sig'))
        if re.fullmatch(r'\d+', str(r['c'][0]).strip()) and re.match(r'^(석사|박사)', str(r['c'][2]).strip())]
ws = openpyxl.load_workbook(OUT).active

def norm_date(v):
    s = str(v or '').strip()
    if not s: return ''
    if '예정' in s:
        m = re.search(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', s)
        if m: return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d} (예정)'
        m = re.search(r'(\d{4})[.\-/](\d{1,2})', s)
        return f'{m.group(1)}-{int(m.group(2)):02d} (예정)' if m else s
    if re.fullmatch(r'\d{5}(\.\d+)?', s):
        return (datetime.date(1899,12,30) + datetime.timedelta(days=round(float(s)))).isoformat()
    t = re.sub(r'[.\-/]+','-',s).strip('-')
    m = re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})', t)
    return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}' if m else s

FIX = {'디저털서울문화예술대학교': '디지털서울문화예술대학교'}
def sch(v):
    s = str(v or '').strip(); return FIX.get(s, s)

bad = 0
for i, c in enumerate(rows):
    r = 3 + i
    got = [ws.cell(r, k).value for k in range(1, 24)]
    dept_raw = str(c[5]).strip()
    tm = re.search(r'(\d)\s*급', str(c[28] or ''))
    m = re.fullmatch(r'\s*(\d{2})(\d{2})(\d{2})\s*-\s*(\d)\d{6}\s*', str(c[9] or ''))
    bd = f'{(2000 if m.group(4) in "78" else 1900)+int(m.group(1))}-{m.group(2)}-{m.group(3)}' if m else ''
    exp = [i+1, str(c[2]).strip(), str(c[3]).strip(), re.sub(r'\s+',' ',str(c[4]).strip()),
           re.sub(r'\(.*?\)','',dept_raw).strip(), '중국어' if '이중언어' in dept_raw else '한국어',
           f'{tm.group(1)}급' if tm else '해당없음', str(c[6]).strip(), str(c[7]).strip(), bd,
           sch(c[10]) or None, str(c[11] or '').strip() or None,
           (int(c[12]) if re.fullmatch(r'\d+', str(c[12]).strip()) else (str(c[12]).strip() or None)),
           str(c[13] or '').strip() or None, norm_date(c[14]) or None,
           sch(c[16]) or None, str(c[17] or '').strip() or None, str(c[19] or '').strip() or None, norm_date(c[20]) or None,
           sch(c[22]) or None, str(c[23] or '').strip() or None, str(c[25] or '').strip() or None, norm_date(c[26]) or None]
    diff = [(k, e, g) for k, e, g in zip(range(1,24), exp, got) if (e or None) != (g or None)]
    if diff:
        bad += 1
        print(f'❌ 행{r} {c[3]}: ' + ', '.join(f'{openpyxl.utils.get_column_letter(k)}열 원본={e!r} 생성={g!r}' for k,e,g in diff))
    else:
        print(f'✅ 행{r} {str(c[3]).strip()}')
print(f'\n═══ 결과: 일치 {len(rows)-bad} / {len(rows)} ═══')
