# -*- coding: utf-8 -*-
"""2차 지원자 현황 엑셀 → '학과별 지원자 현황' 양식으로 정리.

기존 배포 양식((박사)간호학과_12명.xlsx)을 **파일째 복제**해 데이터만 교체하므로
헤더 병합·테두리(블록 경계 double)·열너비·틀고정(E3)·필터·인쇄설정이 100% 동일하다.

양식 열 구성:
  A순번 B과정 C이름 D영문명 E지원학과 F트랙 G TOPIK H국적 I성별 J생년월일
  K~O 전문학사(학교명·성적·연제·전공·졸업일자)
  P~S 학사(학교명·성적·전공·졸업일자)
  T~W 석사(학교명·성적·전공·졸업일자)
"""
import json, os, re, shutil, datetime
import openpyxl

SRC_ROWS = r'C:\projects\graduate application\autobot\rows2.json'
TPL = (r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기\지원서류\학과 평가'
       r'\학과별 배포\학과별 지원자 현황\(박사)간호학과_12명.xlsx')
OUT_DIR = r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기'

# 엑셀 오탈자 교정
SCHOOL_FIX = {'디저털서울문화예술대학교': '디지털서울문화예술대학교'}

C = dict(no=0, gubun=2, name=3, en=4, dept=5, nati=6, sex=7, reg=9,
         col=10, uni=16, ma=22, topik=28)


def ymd(v):
    """엑셀 직렬번호/여러 표기 → 'YYYY-MM-DD'. '예정'은 표기 유지."""
    s = str(v or '').strip()
    if not s:
        return ''
    if '예정' in s:
        m = re.search(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', s)
        if m:
            return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d} (예정)'
        m = re.search(r'(\d{4})[.\-/](\d{1,2})', s)
        if m:
            return f'{m.group(1)}-{int(m.group(2)):02d} (예정)'
        return s
    if re.fullmatch(r'\d{5}(\.\d+)?', s):
        d = datetime.date(1899, 12, 30) + datetime.timedelta(days=round(float(s)))
        return d.isoformat()
    t = re.sub(r'[.\-/]+', '-', s).strip('-')
    m = re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})', t)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return s


def birth(reg):
    """등록번호 'YYMMDD-XXXXXXX' → 생년월일. 뒷자리 첫 숫자로 세기 판정(5·6=19xx, 7·8=20xx)."""
    m = re.fullmatch(r'\s*(\d{2})(\d{2})(\d{2})\s*-\s*(\d)\d{6}\s*', str(reg or ''))
    if not m:
        return ''
    yy, mm, dd, code = m.groups()
    cent = 2000 if code in '78' else 1900
    return f'{cent + int(yy)}-{mm}-{dd}'


def school(v):
    s = str(v or '').strip()
    return SCHOOL_FIX.get(s, s)


def yeonje(v):
    s = str(v or '').strip()
    return int(s) if re.fullmatch(r'\d+', s) else (s or None)


rows = json.load(open(SRC_ROWS, encoding='utf-8-sig'))
recs = []
for row in rows:
    c = row['c']
    if not re.fullmatch(r'\d+', str(c[C['no']]).strip()):
        continue
    gubun = str(c[C['gubun']]).strip()
    if not re.match(r'^(석사|박사)', gubun):
        continue
    dept_raw = str(c[C['dept']]).strip()
    topik = str(c[C['topik']] or '').strip()
    tm = re.search(r'(\d)\s*급', topik)
    recs.append([
        None,                                        # A 순번(뒤에서 부여)
        gubun,                                       # B 과정
        str(c[C['name']]).strip(),                   # C 이름
        re.sub(r'\s+', ' ', str(c[C['en']]).strip()),  # D 영문명
        re.sub(r'\(.*?\)', '', dept_raw).strip(),    # E 지원학과
        '중국어' if '이중언어' in dept_raw else '한국어',  # F 트랙
        f'{tm.group(1)}급' if tm else '해당없음',      # G TOPIK
        str(c[C['nati']]).strip(),                   # H 국적
        str(c[C['sex']]).strip(),                    # I 성별
        birth(c[C['reg']]),                          # J 생년월일
        school(c[C['col']]), str(c[C['col'] + 1] or '').strip() or None,
        yeonje(c[C['col'] + 2]), str(c[C['col'] + 3] or '').strip() or None,
        ymd(c[C['col'] + 4]) or None,                # K~O 전문학사
        school(c[C['uni']]) or None, str(c[C['uni'] + 1] or '').strip() or None,
        str(c[C['uni'] + 3] or '').strip() or None, ymd(c[C['uni'] + 4]) or None,   # P~S 학사
        school(c[C['ma']]) or None, str(c[C['ma'] + 1] or '').strip() or None,
        str(c[C['ma'] + 3] or '').strip() or None, ymd(c[C['ma'] + 4]) or None,     # T~W 석사
    ])

for i, r in enumerate(recs, 1):
    r[0] = i
    if not r[10]:
        r[10] = None       # 전문학사 없음 → 빈칸

out = os.path.join(OUT_DIR, f'(2차)2026학년도 후기 대학원 지원자 현황_{len(recs)}명.xlsx')
shutil.copyfile(TPL, out)
wb = openpyxl.load_workbook(out)
ws = wb.active

FIRST, LAST_TPL = 3, ws.max_row          # 템플릿 데이터 행 범위
for i, rec in enumerate(recs):
    r = FIRST + i
    for j, v in enumerate(rec):
        ws.cell(r, j + 1).value = v

last = FIRST + len(recs) - 1
if LAST_TPL > last:                       # 남는 템플릿 행 삭제
    ws.delete_rows(last + 1, LAST_TPL - last)
ws.auto_filter.ref = f'A2:W{last}'

wb.save(out)
print('생성:', out)
print('행 수:', len(recs))
for rec in recs:
    print(f'  {rec[0]:>2} {rec[1]:<4} {rec[2]:<5} {rec[4]:<14} {rec[5]} {rec[6]:<5} {rec[9]}'
          f' | 전문 {rec[10] or "-"} | 학사 {rec[15] or "-"} {rec[18] or ""} | 석사 {rec[19] or "-"} {rec[22] or ""}')
