# -*- coding: utf-8 -*-
"""통장 입금자명(영문, 흔히 10자로 잘림)을 대학원 지원자 명단과 대조한다.

사용:
  python match_payer.py "ZHOUZHONGL" [더 있으면 나열...]
  python match_payer.py --tsv 입금내역.tsv        # 입금자명<TAB>금액
결과: 한글성명 <TAB> 영문성명 <TAB> 과정 <TAB> 지원학과 <TAB> 금액
"""
import io, os, re, sys
import openpyxl

# 지원자 정보를 가진 파일들 (앞에 있는 것부터 찾는다)
SOURCES = [
    (r'C:\Users\user\Desktop\★국제협력처\0. 유학생 통계\2026-2학기 학위 과정 유학생 현황'
     r'\★20260901 기준 외국인 유학생 재학생 통계(학번 기재).xlsx',
     '전형료 입금확인', {'ko': '한글명', 'en': '영문명', 'course': '과정',
                        'dept': '학부(과)', 'kind': '학적구분'}),
    (r'C:\Users\user\Desktop\★국제협력처\2. 대학원\2026학년도 후기\전형료'
     r'\2026학년도 후기 대학원 지원자 전형료 납부현황(39명).xlsx',
     '전형료 납부현황', {'ko': '이름', 'en': '영문명', 'course': '과정', 'dept': '지원학과'}),
]

norm = lambda s: re.sub(r'[^A-Z0-9]', '', str(s or '').upper())


def load():
    people = []
    for path, sheet, cols in SOURCES:
        if not os.path.exists(path):
            continue
        ws = openpyxl.load_workbook(path, data_only=True)[sheet]
        H = {str(ws.cell(1, c).value).strip(): c
             for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        for r in range(2, ws.max_row + 1):
            ko = ws.cell(r, H[cols['ko']]).value if cols['ko'] in H else None
            if not ko:
                continue
            en = ws.cell(r, H[cols['en']]).value if cols['en'] in H else ''
            course = str(ws.cell(r, H[cols['course']]).value or '') if cols['course'] in H else ''
            dept = str(ws.cell(r, H[cols['dept']]).value or '') if cols['dept'] in H else ''
            kind = str(ws.cell(r, H[cols['kind']]).value or '') if cols.get('kind') in H else ''
            # 수입처리 서식의 '과정' 표기 : 석사-신입 / 박사-신입 / 석사-편입
            k = '편입' if '편입' in kind else '신입'
            people.append({'ko': str(ko).strip(), 'en': str(en or '').strip(),
                           'course': ('%s-%s' % (course.strip(), k)) if course.strip() else '',
                           'dept': dept.strip(), 'src': os.path.basename(path)})
    return people


def find(payer, people):
    """통장 표기는 잘리므로 '앞부분이 같으면' 같은 사람으로 본다."""
    p = norm(payer)
    if len(p) < 5:
        return []
    hit = [x for x in people if norm(x['en']).startswith(p) or p.startswith(norm(x['en']))]
    # 같은 사람이 두 자료에 있으면 하나로
    seen, uniq = set(), []
    for x in hit:
        k = (x['ko'], x['dept'])
        if k not in seen:
            seen.add(k)
            uniq.append(x)
    return uniq


def main():
    people = load()
    print('지원자 자료 %d건 적재' % len(people), file=sys.stderr)

    pays = []
    if len(sys.argv) > 2 and sys.argv[1] == '--tsv':
        for ln in io.open(sys.argv[2], encoding='utf-8'):
            if ln.strip():
                f = ln.rstrip('\n').split('\t')
                pays.append((f[0], f[1] if len(f) > 1 else '50000'))
    else:
        pays = [(a, '50000') for a in sys.argv[1:]]

    out, bad = [], []
    for payer, amt in pays:
        got = find(payer, people)
        if len(got) == 1:
            g = got[0]
            out.append('%s\t%s\t%s\t%s\t%s' % (g['ko'], g['en'], g['course'], g['dept'],
                                               str(amt).replace(',', '')))
        else:
            bad.append((payer, got))

    for ln in out:
        print(ln)
    for payer, got in bad:
        print('## 확인필요 %s → 후보 %d명 %s' % (payer, len(got),
              ', '.join('%s(%s)' % (g['ko'], g['dept']) for g in got)), file=sys.stderr)
    if bad:
        sys.exit(2)


if __name__ == '__main__':
    main()
